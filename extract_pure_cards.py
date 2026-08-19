import openpyxl
import json
import datetime
from dateutil.relativedelta import relativedelta

def format_date(val):
    if val is None:
        return None
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime('%d/%m/%Y')
    s = str(val).strip()
    if s in ['', '-', 'None', '?', '6-24']:
        return None
    for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%y', '%d-%m-%y'):
        try:
            d = datetime.datetime.strptime(s, fmt)
            return d.strftime('%d/%m/%Y')
        except ValueError:
            pass
    return None

def add_6_months(date_str):
    if not date_str:
        return None
    try:
        d = datetime.datetime.strptime(date_str, '%d/%m/%Y')
        exp = d + relativedelta(months=6)
        return exp.strftime('%d/%m/%Y')
    except Exception:
        return None

def clean_str(val):
    if val is None:
        return None
    s = str(val).strip()
    if s in ['', 'None', '-']:
        return None
    return s

def clean_bool(val):
    if val is None:
        return None
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        if val == 1: return True
        if val == 0: return False
    s = str(val).strip().lower()
    if s in ['true', '1', 'si', 'sí', 's', 'verdadero']: return True
    if s in ['false', '0', 'no', 'n', 'falso']: return False
    return None

def clean_importe(val, obs):
    if val is None:
        return None, obs
    if isinstance(val, (int, float)):
        return float(val), obs
    s_imp = str(val).strip()
    if s_imp in ['', '-', '?']:
        return None, obs
    try:
        return float(s_imp.replace('€', '').replace(' ', '').replace(',', '.').strip()), obs
    except ValueError:
        if not obs:
            obs = s_imp
        else:
            obs = f'{obs} | Importe: {s_imp}'
        return None, obs

wb = openpyxl.load_workbook('tarjetas_regalo/OPARI TXARTELAK.xlsx', data_only=True)

all_cards = []
auto_id = 1

# 1. OT PERSONALIZADAS
ws_p = wb['OT PERSONALIZADAS']
for r in range(3, ws_p.max_row + 1):
    vals = [ws_p.cell(r, c).value for c in range(1, 14)]
    name = clean_str(vals[1])
    if not name or name.lower() in ['nombre', 'datos comprador']:
        continue
    
    id_val = auto_id
    auto_id += 1
    
    telf = clean_str(vals[2])
    nombre_comensal = clean_str(vals[3])
    codigo = clean_str(vals[4])
    obs = clean_str(vals[6])
    importe, obs = clean_importe(vals[5], obs)
    entregado = clean_bool(vals[7])
    fecha_entrega = format_date(vals[8])
    pagado = clean_bool(vals[9])
    fecha_pago = format_date(vals[10])
    usado = clean_bool(vals[11])
    fecha_cad = format_date(vals[12])
    
    if not fecha_cad:
        base_date = fecha_pago or fecha_entrega
        if base_date:
            fecha_cad = add_6_months(base_date)
            
    all_cards.append({
        'id': id_val,
        'origen_pestana': 'OT PERSONALIZADAS',
        'fila_excel': r,
        'nombre_compra': name,
        'nombre_comensal': nombre_comensal,
        'telefono_compra': telf,
        'codigo_tarjeta_regalo': codigo,
        'importe': importe,
        'observaciones': obs,
        'creada_en_revo': None,
        'fecha_compra': None,
        'entregado': entregado,
        'fecha_entrega': fecha_entrega,
        'pagado': pagado,
        'fecha_pago': fecha_pago,
        'usado': usado,
        'fecha_caducidad': fecha_cad
    })

# 2. OT WIX
ws_w = wb['OT WIX']
for r in range(3, ws_w.max_row + 1):
    vals = [ws_w.cell(r, c).value for c in range(1, 12)]
    name = clean_str(vals[1])
    if not name or name.lower() in ['nombre', 'datos comprador']:
        continue
    
    id_val = auto_id
    auto_id += 1
    
    telf = clean_str(vals[2])
    codigo = clean_str(vals[4])
    obs = clean_str(vals[6])
    importe, obs = clean_importe(vals[5], obs)
    fecha_compra = format_date(vals[7])
    creada_en_revo = clean_bool(vals[8])
    usado = clean_bool(vals[9])
    fecha_cad = format_date(vals[10])
    
    if not fecha_cad and fecha_compra:
        fecha_cad = add_6_months(fecha_compra)
        
    all_cards.append({
        'id': id_val,
        'origen_pestana': 'OT WIX',
        'fila_excel': r,
        'nombre_compra': name,
        'nombre_comensal': None,
        'telefono_compra': telf,
        'codigo_tarjeta_regalo': codigo,
        'importe': importe,
        'observaciones': obs,
        'creada_en_revo': creada_en_revo,
        'fecha_compra': fecha_compra,
        'entregado': None,
        'fecha_entrega': None,
        'pagado': None,
        'fecha_pago': None,
        'usado': usado,
        'fecha_caducidad': fecha_cad
    })

# 3. OT SHOPIFY
ws_s = wb['OT SHOPIFY']
for r in range(3, ws_s.max_row + 1):
    vals = [ws_s.cell(r, c).value for c in range(1, 10)]
    name = clean_str(vals[1])
    if not name or name.lower() in ['nombre', 'datos comprador']:
        continue
    
    id_val = auto_id
    auto_id += 1
    
    codigo = clean_str(vals[2])
    obs = clean_str(vals[4])
    importe, obs = clean_importe(vals[3], obs)
    fecha_compra = format_date(vals[5])
    creada_en_revo = clean_bool(vals[6])
    usado = clean_bool(vals[7])
    fecha_cad = format_date(vals[8])
    
    if not fecha_cad and fecha_compra:
        fecha_cad = add_6_months(fecha_compra)
        
    all_cards.append({
        'id': id_val,
        'origen_pestana': 'OT SHOPIFY',
        'fila_excel': r,
        'nombre_compra': name,
        'nombre_comensal': None,
        'telefono_compra': None,
        'codigo_tarjeta_regalo': codigo,
        'importe': importe,
        'observaciones': obs,
        'creada_en_revo': creada_en_revo,
        'fecha_compra': fecha_compra,
        'entregado': None,
        'fecha_entrega': None,
        'pagado': None,
        'fecha_pago': None,
        'usado': usado,
        'fecha_caducidad': fecha_cad
    })

print(f"Total procesadas con exactitud absoluta: {len(all_cards)}")
with open('tarjetas_regalo/tarjetas_regalo_unificadas.json', 'w', encoding='utf-8') as f:
    json.dump(all_cards, f, ensure_ascii=False, indent=2)

print("Archivo tarjetas_regalo/tarjetas_regalo_unificadas.json actualizado con éxito.")
